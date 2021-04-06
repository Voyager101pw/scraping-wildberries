import os.path
import re
import time
import random
import tkinter.messagebox
from datetime import datetime
# Скачанные библиотеки
import openpyxl
import pickle
import requests
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
import telebot
import logging

# class Analytics_sheet:
#     def check_old_price:
def get_html():
    global my_cookies
    my_cookies = open_cookie()
    while cookies_is_right():
        my_cookies = create_cookies()
    return requests.get(parsing_link, cookies=my_cookies).text


def open_cookie():
    while True:
        if file_is_exist(cookie_path):
            return load_cookies()
        else:
            return create_cookies()


def load_cookies():
    # Функция загружает куки и преобразует их из типа [] > {} - что бы requests.get мог их прочитать
    cookies_list = pickle.load(open(cookie_path, 'rb'))
    cookies = {}
    for cookie in cookies_list:
        cookies[cookie['name']] = cookie['value']
    return cookies


def create_cookies():
    tkinter.messagebox.showerror('Ошибка Cookie!', 'Cookie файлы не существуют, либо не подходят!\n'
                                                   'Для создания новых cookie необходимо авторизоватся вручную!\n''Вам будет выделено 30 секунд!')
    driver = open_driver()
    driver.get('https://www.wildberries.ru/security/login')
    time.sleep(30)
    pickle.dump(driver.get_cookies(), open(cookie_path, 'wb'))
    driver.quit()
    return load_cookies()


def open_driver():
    option = webdriver.FirefoxOptions()
    option.set_preference('dom.webdriver.enabled', False)  # Отключаем у браузера отображение Вебдрайвера
    option.set_preference('general.useragent.override', \
                          'Mozilla/5.0 (Windows NT 6.1; WOW64;\
                           Trident/7.0; AS; rv:11.0) like Gecko')  # Меням юзерагент
    driver = webdriver.Firefox(options=option)
    return driver


def cookies_is_right():
    html = requests.get(parsing_link, cookies=my_cookies).text
    # Если данные о товаре имеются на странице, то куки верные.
    if '"goodsName":' in html and '"brandName":' in html or '"priceWithCoupon":' in html:
        return False  # Останавливаем цикл
    return True  # Цикл продолжает работать


def get_data_from_html(html):
    data = [[], [], [], []]  # [[article], [brands],[names],[prices]]
    html = html.replace('{', '\n').replace('}', '').split('\n')
    for line in html:
        if '"cod1S":' in line:
            [data[0].append(re.sub('.*:', '', i)) for i in re.findall('"cod1S":[0-9]*', str(line))]

            [data[1].append(i.replace('"', '').replace('brandName:', '')) \
             for i in re.findall('"brandName":"[a-zA-Zа-яА-Я0-9 .,-]*', line)]  # Бренд

            [data[2].append(i.replace('\\"', '"')) \
             for i in re.findall(f'(?<="goodsName":").*?(?=",)', str(line))]  # Название

            [data[3].append(re.sub(r'[a-zA-Z]*\W', '', i)) if re.sub(r'[a-zA-Z]*\W', '', i) != '0' else data[3].append(
                'Нет на складе.') \
             for i in re.findall('"priceWithCouponAndDiscount":[0-9]*', line)]  # Цена
    return data


def file_is_exist(path):
    # Проверяет путь до куков, и возвращает bool
    return os.path.exists(path)


# _______________________OPEN_PY_EXCEL__________________________
def pyexcel(data):
    book = create_excel_book(data)  # Книга
    book, sheet0, sheet1 = removing_and_add_goods(book, book[book.sheetnames[0]], book[book.sheetnames[1]], data)
    check_update_price(book, book[book.sheetnames[0]], data)


def create_excel_book(data):
    if not os.path.exists(excel_path):
        excel_book = openpyxl.Workbook()
        excel_book[excel_book.sheetnames[0]].title = 'Отложенные товары'
        excel_book.create_sheet('Min-Max цен')
        time_now = str(datetime.now().strftime("%H:%M | %d.%m.%Y"))
        titles = [{0: 'Артикул', 1: 'Название бренда', 2: 'Название товара', 3: time_now},
                  {0: 'Артикул', 1: 'Бренд', 2: 'Название', 3: 'Now (Min | Max)', 4: 'Date min | Date max'}]
        widths = [{0: 11, 1: 18, 2: 50, 3: 20},
                  {0: 11, 1: 18, 2: 50, 3: 25, 4: 25}]

        # Создание структуры отображения данных на двух страницах книги.
        for repeat in range(2):  # repeat = 1стр и 2стр
            sheet = excel_book[excel_book.sheetnames[repeat]]
            for i in range(len(data)):
                sheet[get_column_letter(i + 1) + str(1)].value = titles[0][i]
                cell_styles(sheet, i + 1, 1, True, widths[0][i], None)
                for j in range(len(data[i])):
                    if repeat == 0:  # Инструкции для 1 СТРАНИЦЫ
                        sheet[get_column_letter(i + 1) + str(j + 2)].value = data[i][j]
                    else:  # Инструкции для 2 СТРАНИЦЫ
                        if i < 3:  # Первые 3 столбца собираются как в первой странице книги
                            sheet[get_column_letter(i + 1) + str(j + 2)].value = data[i][j]
                        else:  # Остальные столбцы собираются кастомно
                            cell_styles(sheet, i + 2, j + 2, True, None, None)
                            if data[i][j] == 'Нет на складе.':
                                sheet[get_column_letter(i + 1) + str(j + 2)].value = f'0 (0 | 0)'
                            else:
                                sheet[get_column_letter(i + 1) + str(j + 2)].value \
                                    = f'{data[i][j]} ({data[i][j]} | {data[i][j]})'

                            sheet[get_column_letter(i + 2) + str(j + 2)].value = \
                                f'{datetime.now().strftime("%d.%m.%Y")} | {datetime.now().strftime("%d.%m.%Y")}'

                        if i == 3 and j == len(data) - 1:
                            sheet[get_column_letter(i + 2) + str(1)].value = titles[0][i + 1]
                            cell_styles(sheet, i + 2, 1, True, widths[0][i + 1], None)

                    if i == 3:
                        cell_styles(sheet, i + 1, j + 2, True, None, None)
            titles.pop(0)
            widths.pop(0)
        excel_book.save(excel_path)
    return openpyxl.load_workbook(excel_path)


def removing_and_add_goods(book, sheet0, sheet1, new_goods):
    new_articles = new_goods[0]  # Список новых артикулов
    if sheets_is_not_empty(sheet0, sheet1):
        articles = sheet0['A'][1:sheet0.max_row]
        all_old_article = [article.value for article in articles]
        # ЦИКЛ 1 - Удаление товаров которых нет в запросе \ на сайте > Отложенные товары
        row_int = 2  # Начинать будем со 2 строчки, поскольку строка 1 содержит шапку
        len_rows = sheet0.max_row  # Общее количество строк на странице, включая шапку
        while row_int <= len_rows:
            if sheet0['A' + str(row_int)].value not in new_articles and sheet0['A' + str(row_int)].value is not None:
                print(f'\t[-] Товар удален: {sheet0["C" + str(row_int)].value} [{sheet0["A" + str(row_int)].value}]')
                send_in_telegram(f'\t[-] Товар удален: {sheet0["C" + str(row_int)].value} [{sheet0["A" + str(row_int)].value}]')
                sheet0.delete_rows(row_int)  # Если строка удалена, следующая строка падает на координаты удаленной
                sheet1.delete_rows(row_int)  # строки, по этому цикл заново повторяется без увеличение индекса row
                continue  # что бы заново проверить товар в координатах ранее удаленного товара  A2 > A2
            row_int += 1  # A2 > A3 > A4 ...

        # ЦИКЛ 2 - Добавление и Перемещение товаров согласно порядку на сайте
        for i in range(len(new_articles)):
            # ДОБАВЛЕНИЕ ТОВАРОВ
            if new_articles[i] != sheet0[f'A{i + 2}'].value and new_articles[i] not in all_old_article:
                print(f'\t[+] Обнаружен товар: {new_goods[2][i]} [{new_articles[i]}]')
                send_in_telegram(f'\t[+] Обнаружен товар: {new_goods[2][i]} [{new_articles[i]}]')
                sheet0.insert_rows(i + 2, 1)
                sheet1.insert_rows(i + 2, 1)
                for j in range(5):
                    column = get_column_letter(j + 1)
                    if j <= 2:  # Столбцы > A,B,C
                        sheet0[f'{column + str(i + 2)}'].value = new_goods[j][i]
                        sheet1[f'{column + str(i + 2)}'].value = new_goods[j][i]
                    if j == 3 and new_goods[3][i] != 'Нет на складе.' and new_goods[3][i] is not None:  # Столбец > D
                        sheet0[f'D{i + 2}'].value = new_goods[3][i]
                        sheet1[f'D{i + 2}'].value = f'{new_goods[3][i]} ({new_goods[3][i]} | {new_goods[3][i]})'
                    if new_goods[3][i] == 'Нет на складе.' or new_goods[3][i] is None:
                        sheet0[f'D{i + 2}'].value = 'Нет на складе.'
                        sheet1[f'D{i + 2}'].value = f'0 (0 | 0)'
                    if j == 4:
                        sheet1[column + str(i + 2)].value = \
                            f'{datetime.now().strftime("%d.%m.%Y")} | {datetime.now().strftime("%d.%m.%Y")}'
                cell_styles(sheet0, 4, i + 2, True, None, 'ffd700')
                cell_styles(sheet1, 4, i + 2, True, None, None)
                cell_styles(sheet1, 5, i + 2, True, None, None)

            # ПЕРЕМЕЩЕНИЕ ТОВАРОВ
            if new_articles[i] != sheet0[f'A{i + 2}'].value and new_articles[i] in all_old_article:  # Перемещение
                print(f'\t[>] Товар перемещен: {new_goods[2][i]} [{new_goods[0][i]}] ')
                sheet0.insert_rows(i + 2, 1)
                sheet1.insert_rows(i + 2, 1)
                rewrite_cell(i, sheet0, sheet1, new_articles)
            book.save(excel_path)
        return openpyxl.load_workbook(excel_path), sheet0, sheet1


def rewrite_cell(i, sheet0, sheet1, new_articles):
    row_old_article = None
    for old_article in sheet0['A'][1:sheet0.max_row]:  # old_article= cell(A[2]) > cell(A[end_row])
        if old_article.value == new_articles[i]:
            row_old_article = int(old_article.coordinate[1:])  # КООРДИНАТЫ: A2 > 2
            break
    if row_old_article is not None:  # Если строка старого артикула найдена, то есть не пуста / not None
        for column_int in range(1, sheet0.max_column + 1):  # A,B,C,D,... max_column + 1 = END_COLUMN
            if sheet0[f'{get_column_letter(column_int)}{row_old_article}'].value is not None:
                # Записываем в освобожденную стр
                old_cell = sheet0[f'{get_column_letter(column_int)}{row_old_article}'].value
                sheet0[f'{get_column_letter(column_int)}{i + 2}'].value = old_cell

                # Выравнивание по центру для всех столбцов с ценами от столбца D до конца
                if column_int > 3 and '+' in old_cell:
                    cell_styles(sheet0, column_int, i + 2, True, None, 'ff0000')
                elif column_int > 3 and '-' in old_cell:
                    cell_styles(sheet0, column_int, i + 2, True, None, '85bb65')
                elif column_int > 3 and '+' not in old_cell and '-' not in old_cell:
                    cell_styles(sheet0, column_int, i + 2, True, None, 'cccccc')
            else:  # Если значение ячейки пустое/None,
                break

        for column_int in range(1, sheet1.max_column + 1):  # СТРАНИЦА 2 - ПЕРМЕЩЕНИЕ ячеек
            sheet1[f'{get_column_letter(column_int)}{i + 2}'].value \
                = sheet1[f'{get_column_letter(column_int)}{row_old_article}'].value
        cell_styles(sheet1, 4, i + 2, True, None, None)
        cell_styles(sheet1, 5, i + 2, True, None, None)
        sheet0.delete_rows(row_old_article)  # Удаление старого расположения
        sheet1.delete_rows(row_old_article)  # Удаление старого расположения


def sheets_is_not_empty(sheet0, sheet1):
    if sheet0.max_column > 1 and sheet0.max_row > 1 and sheet1.max_column > 1 and sheet1.max_row > 1:
        return True
    print('Таблицы пустые!')
    exit()


def cell_styles(sheet, col, row, alignment, width_value, color_code):
    if alignment:
        sheet[get_column_letter(col) + str(row)].alignment = Alignment(horizontal='center', vertical='center')
    if width_value is not None:
        sheet.column_dimensions[get_column_letter(col)].width = width_value
    if color_code is not None:
        sheet[get_column_letter(col) + str(row)].fill = PatternFill(fill_type='solid', start_color=color_code)


def check_update_price(book, sheet0, data):
    if price_updated(sheet0, data):
        sheet0.move_range(f'D1:{get_column_letter(sheet0.max_column)}{sheet0.max_row}', cols=1)
        cell_styles(sheet0, sheet0.max_column, None, False, 18, None)  # Делаем ширину столба для устаревших цен
        print(f'[*] Цены обновлены : {datetime.now().strftime("%H:%M:%S")}')
        send_in_telegram(f'[*] Цены обновлены : {datetime.now().strftime("%H:%M:%S")}')
        # Цикл НАПОЛНЕНИЯ НОВЫХ ЦЕНЫ
        for i in range(len(data[0])):
            add_price_sheet0(book, data, i)  # Попутно функция будет изменять цены в sheet1
        cell_styles(sheet0, 4, 1, True, 18, None)
        sheet0['D1'].value = datetime.now().strftime("%H:%M | %d.%m.%Y")  # Дата последнего обновления цен
        book.save(excel_path)


def price_updated(sheet, data):  # ПРОВЕРЯЕТ ЦЕНУ на обновление
    index_last_row = sheet.max_row
    for row in range(2, index_last_row + 1):  # со 2 строки до крайней строки вкл.(index_last_row + 1), 1 строка - шапка
        if sheet['D' + str(row)].value != 'Нет на складе.' and sheet['D' + str(row)].value is not None:
            price_now = re.search('^\d*', sheet['D' + str(row)].value).group()
            if price_now != data[3][row-2]:  # Проверка на изменение цены
                return True  # Цена обновлена
        if sheet['D' + str(row)].value == 'Нет на складе.':
            if sheet['D' + str(row)].value != data[3][row - 2]:  # Проверка на наличие товара
                return True  # Товар появился
    return False  # Цены не изменились, программа перепрыгивает алгоритм и ждет +\- 2ч.


def add_price_sheet0(book, data, i):
    sheet0 = book[book.sheetnames[0]]
    sheet1 = book[book.sheetnames[1]]
    cell_styles(sheet0, 5, i + 2, True, 18, None)
    new_price_str = data[3][i]
    # Нахождение новой и старой цены (регуляркой отделаем цену от других символов)
    if sheet0['E' + str(i + 2)].value != 'Нет на складе.' and sheet0['E' + str(i + 2)].value is not None:
        old_price_str = re.search('^[0-9]*', str(sheet0['E' + str(i + 2)].value)).group()
    else:
        old_price_str = sheet0['E' + str(i + 2)].value

    if new_price_str == old_price_str:  # ЦЕНА НЕ ИЗМЕНИЛАСЬ
        cell_styles(sheet0, 4, i + 2, True, None, 'cccccc')
        sheet0['D' + str(i + 2)].value = old_price_str
        return  # Серая цена
    else:  # ЦЕНА ИЗМЕНИЛАСЬ
        # Если старая или новая цена это не цифры то:
        if new_price_str != 'Нет на складе.' and old_price_str == 'Нет на складе.' or old_price_str is None:
            cell_styles(sheet0, 4, i + 2, True, None, 'ffd700')
            sheet0['D' + str(i + 2)].value = new_price_str
            add_price_sheet1(sheet1, new_price_str, i)
            send_in_telegram(f'\t[+]Товар появился: {data[2][i]}')
            return print(f'\t[+]Товар появился: {data[2][i]}')  # Товар появился - желтый ценник
        elif new_price_str == 'Нет на складе.' and old_price_str != 'Нет на складе.':
            cell_styles(sheet0, 4, i + 2, True, None, 'AF6666')
            sheet0['D' + str(i + 2)].value = new_price_str
            add_price_sheet1(sheet1, new_price_str, i)
            send_in_telegram(f'\t[-]Товар пропал с продажи: {data[2][i]}')
            return print(f'\t[-]Товар пропал с продажи: {data[2][i]}')  # Товар пропал - бордовый ценник
        else:  # Если старая и новая цена это числа
            if int(new_price_str) > int(old_price_str):
                cell_styles(sheet0, 4, i + 2, True, None, 'ff0000')
                sheet0['D' + str(i + 2)].value = f"{new_price_str} | +{int(new_price_str) - int(old_price_str)}р."
                add_price_sheet1(sheet1, new_price_str, i)
                send_in_telegram(f'\t[$]Повышение цены на {int(new_price_str) - int(old_price_str)}р : {data[2][i]}')
                return print(f'\t[$]Повышение цены на {int(new_price_str) - int(old_price_str)}р : {data[2][i]}')
            elif int(new_price_str) < int(old_price_str):
                cell_styles(sheet0, 4, i + 2, True, None, '85bb65')
                sheet0['D' + str(i + 2)].value = f"{new_price_str} | -{int(old_price_str) - int(new_price_str)}р."
                add_price_sheet1(sheet1, new_price_str, i)
                send_in_telegram(f'\t[$]Снижение цены на {int(old_price_str) - int(new_price_str)}р : {data[2][i]}')
                return print(f'\t[$]Снижение цены на {int(old_price_str) - int(new_price_str)}р : {data[2][i]}')


def add_price_sheet1(sheet1, new_price, i):
    if new_price == 'Нет на складе.' or new_price is None:
        sheet1['D' + str(i + 2)].value = '0 ' + str(re.search('\(.*\)', sheet1['D' + str(i + 2)].value).group())
        return
    new_price = int(new_price)
    min_price = int(re.search('(?<=\()\d+', sheet1['D' + str(i + 2)].value).group())
    max_price = int(re.search('\d+(?=\))', sheet1['D' + str(i + 2)].value).group())
    if min_price == 0:
        sheet1['D' + str(i + 2)].value = f'{new_price} ({new_price} | {max_price})'
        return
    elif new_price < min_price:
        date_max = re.search('..\...\.....$', sheet1['E' + str(i + 2)].value).group()
        sheet1['E' + str(i + 2)].value = f'{datetime.now().strftime("%d.%m.%Y")} | {date_max}'
        sheet1['D' + str(i + 2)].value = f'{new_price} ({new_price} | {max_price})'
        return
    elif new_price > max_price:
        date_min = re.search('^..\...\.....', sheet1['E' + str(i + 2)].value).group()
        sheet1['E' + str(i + 2)].value = f'{date_min} | {datetime.now().strftime("%d.%m.%Y")}'
        sheet1['D' + str(i + 2)].value = f'{new_price} ({min_price} | {new_price})'
        return
    sheet1['D' + str(i + 2)].value = f'{new_price} ({min_price} | {max_price})'


def send_in_telegram(message):
    while True:
        try:
            telegram_bot.send_message(713887294, message)
            break
        except Exception as e:
            logging.info(e)
            time.sleep(5)


def main():
    # Инициализируем телеграм бота, указав TOKEN
    global telegram_bot
    telegram_bot = telebot.TeleBot('1724944326:AAESiVu-pIRyYjMlyvILQwXeCeB4AbIHsUw')
    while True:
        tkinter.Tk().withdraw()
        global parsing_link, cookie_path, excel_path, html
        cookie_path = 'cookies.pkl'
        excel_path = 'WildBerries.xlsx'
        parsing_link = 'https://www.wildberries.ru/lk/poned/data?page=1&pageSize=100&group=0'
        html = get_html()
        data = get_data_from_html(html)
        pyexcel(data)
        try:
            random.seed(datetime.now())
            time.sleep(random.randint(5400, 7200))
        except:
            print('---Программа остановлена---')
            exit()


if __name__ == '__main__':
    main()

# with open('data1.txt', 'r') as f:
#     data = [[], [], [], []]
#     iter = 0
#     for line in f:
#         data[iter] = line.replace('\n','').replace("'","").split(',')
#         iter += 1


# sheet1 = book[book.sheetnames[1]]
# sheet0 = book[book.sheetnames[0]]
#
# for i in range(len(sheet0['D'][1:sheet0.max_row])):
#     j = 4
#     max = 0
#     min = 999999
#     while True:
#         try:
#             price = int(re.search('^\d*', sheet0[get_column_letter(j)+str(i+2)].value).group())
#         except:
#             break
#         if price > max:
#             max = price
#         if price < min:
#             min =price
#         if sheet0[get_column_letter(j+1)+str(i+2)].value is None:
#             break
#         j += 1
#     if max == 0 and min == 999999:
#         min = 0
#         max = 0
#     now = int(re.search('^\d*', sheet1['D' + str(i + 2)].value).group())
#     sheet1['D'+ str(i + 2)].value = f'{now} ({min} | {max})'
# book.save('WildBerries.xlsx')
# exit()
